import subprocess
import platform

import pandas as pd
from openpyxl import Workbook
import datetime

def new_excel(file_path):
    wb = Workbook()

    # Active worksheet
    ws = wb.active

    s1 = pd.Series(['11', True, 123]).to_list()
    ws['A1'] = 42
    ws.append([1, 2, 3, 4])
    ws['A2'] = datetime.datetime.now()
    ws.append(s1)

    wb.save(file_path)

def open_excel_file(file_path):
    if platform.system() == "Windows":
        excel_path = r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE"
        subprocess.run([excel_path, file_path])
    elif platform.system() == "Darwin":
       print("Entro1")
       subprocess.run(["open", file_path])
    else:
       subprocess.run(["xdg-open", file_path])
       print("Entro2")